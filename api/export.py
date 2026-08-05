"""Excel export — the aged position plus whatever follow-up work has been recorded."""

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import aging
from db import STATUSES

STATUS_LABEL = dict(STATUSES)

THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill('solid', fgColor='1F3864')
HDR_FONT = Font(bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(bold=True, size=14, color='1F3864')
SUB_FONT = Font(italic=True, size=10, color='595959')
TOT_FILL = PatternFill('solid', fgColor='DDEBF7')
HOT_FILL = PatternFill('solid', fgColor='FCE4E4')
WARN_FILL = PatternFill('solid', fgColor='FFF2CC')
CREDIT_FONT = Font(color='C00000')
OK_FILL = PatternFill('solid', fgColor='E4F3EC')
MONEY = '#,##0.00;[Red]-#,##0.00'
DATE_FMT = 'yyyy-mm-dd'


def _header(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill, c.font, c.border = HDR_FILL, HDR_FONT, BORDER
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 30


def _finish(ws, freeze, filter_ref=None):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = freeze
    if filter_ref:
        ws.auto_filter.ref = filter_ref
    ws.page_setup.orientation = 'landscape'
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1


def build(customers, totals, company_label, currency):
    bands = totals['bands']
    as_of = totals['as_of']
    threshold = totals['threshold']
    wb = Workbook()

    # ---- Customers ----
    scope = totals.get('scope', 'aged')
    all_open = scope == 'all'

    ws = wb.active
    ws.title = 'Customers'
    ws['A1'] = (f'{company_label} — Open Receivables, Full Aging' if all_open
                else f'{company_label} — Receivables {threshold}+ Days Overdue')
    ws['A1'].font = TITLE_FONT
    ws['A2'] = (f'As of {as_of}   |   Currency: {currency}   |   '
                f'{totals["customers"]} customers   |   {totals["documents"]} documents')
    ws['A2'].font = SUB_FONT
    ws.merge_cells('A1:M1')
    ws.merge_cells('A2:M2')

    total_label = 'Total Open' if all_open else f'Total {threshold}+'
    headers = (['#', 'Customer', 'Agency', 'Phone', 'City', 'Area', 'Credit Terms', 'Docs', 'Oldest (days)']
               + bands + [total_label, 'Overdue Portion', 'Total Balance',
                          'Status', 'Owner', 'Promise Date', 'Next Action', 'Notes'])
    widths = ([5, 44, 10, 18, 14, 13, 15, 8, 14] + [16] * len(bands)
              + [18, 18, 18, 17, 16, 14, 14, 8])
    _header(ws, 4, headers, widths)

    r = 5
    for idx, c in enumerate(customers, start=1):
        col = 1
        ws.cell(row=r, column=col, value=idx).alignment = Alignment(horizontal='center'); col += 1
        ws.cell(row=r, column=col, value=c['name']).alignment = Alignment(horizontal='right'); col += 1
        ag = ws.cell(row=r, column=col, value='Agency' if c.get('agency') else '')
        ag.alignment = Alignment(horizontal='center')
        if c.get('agency'):
            ag.fill = WARN_FILL
        col += 1
        ws.cell(row=r, column=col, value=c['phone']); col += 1
        ws.cell(row=r, column=col, value=c['city']); col += 1
        ws.cell(row=r, column=col, value=c.get('area') or '—').alignment = \
            Alignment(horizontal='center'); col += 1
        ws.cell(row=r, column=col, value=c.get('payment_term') or '—').alignment = \
            Alignment(horizontal='center'); col += 1
        ws.cell(row=r, column=col, value=c['aged_docs']).alignment = Alignment(horizontal='center'); col += 1

        # Negative means the oldest item is still inside its terms.
        oc = ws.cell(row=r, column=col,
                     value=c['oldest_days'] if c['oldest_days'] > 0 else 'Not due')
        oc.alignment = Alignment(horizontal='center')
        if c['oldest_days'] >= 546:
            oc.fill = HOT_FILL
        elif c['oldest_days'] >= 365:
            oc.fill = WARN_FILL
        col += 1

        for v in c['buckets']:
            cell = ws.cell(row=r, column=col, value=v)
            cell.number_format = MONEY
            if v < 0:
                cell.font = CREDIT_FONT
            col += 1

        t = ws.cell(row=r, column=col, value=c['aged_total'])
        t.number_format = MONEY
        t.font = Font(bold=True, color='C00000') if c['aged_total'] < 0 else Font(bold=True)
        col += 1

        o = ws.cell(row=r, column=col, value=c['overdue_total'])
        o.number_format = MONEY
        if c['overdue_total'] > 0:
            o.font = Font(color='B3261E')
        col += 1

        # What they actually owe. Zero here against a positive aged figure is an
        # invoice already cancelled by an unapplied credit — nothing to chase.
        tb = ws.cell(row=r, column=col, value=c['total_open'])
        tb.number_format = MONEY
        if c.get('settled'):
            tb.fill = DONE_FILL if 'DONE_FILL' in dir() else OK_FILL
            tb.font = Font(bold=True, color='1C6B45')
        col += 1

        ws.cell(row=r, column=col, value=STATUS_LABEL.get(c['status'], c['status'])).alignment = \
            Alignment(horizontal='center'); col += 1
        ws.cell(row=r, column=col, value=c['owner']); col += 1
        pd_ = ws.cell(row=r, column=col, value=c['promise_date'] or '')
        pd_.number_format = DATE_FMT
        pd_.alignment = Alignment(horizontal='center'); col += 1
        na = ws.cell(row=r, column=col, value=c['next_action_date'] or '')
        na.number_format = DATE_FMT
        na.alignment = Alignment(horizontal='center'); col += 1
        ws.cell(row=r, column=col, value=c['notes']).alignment = Alignment(horizontal='center')

        for j in range(1, len(headers) + 1):
            ws.cell(row=r, column=j).border = BORDER
        r += 1

    lbl = ws.cell(row=r, column=2, value=f'TOTAL — {totals["customers"]} customers')
    lbl.font, lbl.fill = Font(bold=True, size=11), TOT_FILL
    lbl.alignment = Alignment(horizontal='right')
    DOCS_COL, FIRST_BAND_COL = 8, 10
    ws.cell(row=r, column=DOCS_COL, value=totals['documents']).alignment = \
        Alignment(horizontal='center')
    col = FIRST_BAND_COL
    for v in totals['band_totals']:
        cell = ws.cell(row=r, column=col, value=v)
        cell.number_format, cell.font, cell.fill = MONEY, Font(bold=True), TOT_FILL
        col += 1
    t = ws.cell(row=r, column=col, value=totals['aged_total'])
    t.number_format, t.font, t.fill = MONEY, Font(bold=True, size=12), TOT_FILL
    o = ws.cell(row=r, column=col + 1, value=totals['overdue_total'])
    o.number_format, o.font, o.fill = MONEY, Font(bold=True), TOT_FILL
    tb = ws.cell(row=r, column=col + 2, value=totals['total_open'])
    tb.number_format, tb.font, tb.fill = MONEY, Font(bold=True), TOT_FILL
    for j in range(1, len(headers) + 1):
        ws.cell(row=r, column=j).border = BORDER
        if j < FIRST_BAND_COL and j != 2:
            ws.cell(row=r, column=j).fill = TOT_FILL
    _finish(ws, 'A5', f'A4:{get_column_letter(len(headers))}{r - 1}')

    # ---- Document detail ----
    ws2 = wb.create_sheet('Document Detail')
    ws2['A1'] = (f'{company_label} — Line Detail, All Open Items' if all_open
                 else f'{company_label} — Line Detail, {threshold}+ Days Overdue')
    ws2['A1'].font = TITLE_FONT
    ws2['A2'] = f'As of {as_of}   |   Currency: {currency}'
    ws2['A2'].font = SUB_FONT
    ws2.merge_cells('A1:K1')
    ws2.merge_cells('A2:K2')
    _header(ws2, 4,
            ['Customer', 'Document', 'Reference', 'Invoice Date', 'Due Date',
             'Days Overdue', 'Age Band', 'Original', 'Balance Due', 'Journal', 'Status'],
            [42, 20, 38, 13, 13, 13, 12, 15, 15, 18, 17])
    r = 5
    for c in customers:
        for d in c['documents']:
            vals = [c['name'], d['doc'], d['ref'], d['inv_date'], d['due_date'],
                    d['days'] if d['days'] > 0 else 0, d['band'],
                    d['original'], d['residual'], d['journal'],
                    STATUS_LABEL.get(c['status'], c['status'])]
            for j, v in enumerate(vals, start=1):
                cell = ws2.cell(row=r, column=j, value=v)
                cell.border = BORDER
                if j in (4, 5):
                    cell.number_format = DATE_FMT
                if j in (8, 9):
                    cell.number_format = MONEY
                if j in (1, 3):
                    cell.alignment = Alignment(horizontal='right')
                if j in (6, 7, 11):
                    cell.alignment = Alignment(horizontal='center')
            if d['days'] >= 546:
                ws2.cell(row=r, column=6).fill = HOT_FILL
                ws2.cell(row=r, column=7).fill = HOT_FILL
            elif d['days'] >= 365:
                ws2.cell(row=r, column=6).fill = WARN_FILL
                ws2.cell(row=r, column=7).fill = WARN_FILL
            if d['residual'] < 0:
                ws2.cell(row=r, column=9).font = CREDIT_FONT
            r += 1
    lbl = ws2.cell(row=r, column=8, value='TOTAL')
    lbl.font, lbl.fill = Font(bold=True, size=11), TOT_FILL
    lbl.alignment = Alignment(horizontal='right')
    t = ws2.cell(row=r, column=9, value=totals['aged_total'])
    t.number_format, t.font, t.fill = MONEY, Font(bold=True, size=12), TOT_FILL
    for j in range(1, 12):
        ws2.cell(row=r, column=j).border = BORDER
    _finish(ws2, 'A5', f'A4:K{r - 1}')

    return wb


def to_bytes(customers, totals, company_label, currency):
    buf = io.BytesIO()
    build(customers, totals, company_label, currency).save(buf)
    return buf.getvalue()


def notes_sheet(wb, conn):
    """Append the full contact log so an exported file carries the history with it."""
    rows = conn.execute(
        'SELECT n.created_at, n.author, n.body, c.name'
        '  FROM notes n LEFT JOIN customers c ON c.partner_id = n.partner_id'
        ' ORDER BY n.created_at DESC'
    ).fetchall()
    if not rows:
        return wb
    ws = wb.create_sheet('Contact Log')
    ws['A1'] = 'Contact Log'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:D1')
    _header(ws, 3, ['When', 'Customer', 'By', 'Note'], [20, 40, 16, 90])
    r = 4
    for row in rows:
        for j, v in enumerate([row['created_at'], row['name'] or '', row['author'] or '',
                               row['body']], start=1):
            cell = ws.cell(row=r, column=j, value=v)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='right' if j in (2, 4) else 'left',
                                       vertical='top', wrap_text=(j == 4))
        r += 1
    _finish(ws, 'A4', f'A3:D{r - 1}')
    return wb


# ------------------------------------------------------------------ collections

def collections_workbook(conn, filters, company_label, currency):
    """Daily collection achievement, by salesperson and by customer."""
    import collections_data

    tot = collections_data.totals(conn, filters)
    wb = Workbook()

    def titles(ws, title, span):
        ws['A1'] = title
        ws['A1'].font = TITLE_FONT
        rng = f"{tot['first_date'] or ''} to {tot['last_date'] or ''}"
        ws['A2'] = (f'{rng}   |   Currency: {currency}   |   '
                    f"{tot['receipts']} receipts over {tot['days']} days")
        ws['A2'].font = SUB_FONT
        ws.merge_cells(f'A1:{get_column_letter(span)}1')
        ws.merge_cells(f'A2:{get_column_letter(span)}2')

    # ---- Summary ----
    ws = wb.active
    ws.title = 'Summary'
    titles(ws, f'{company_label} — Collections', 4)
    _header(ws, 4, ['Measure', 'Value'], [34, 22])
    r = 5
    for name, value, fmt in [
        ('Total collected', tot['total'], MONEY),
        ('Applied to invoices', tot['applied'], MONEY),
        ('Received but unapplied', tot['unapplied'], MONEY),
        ('Receipts', tot['receipts'], '#,##0'),
        ('Customers who paid', tot['customers'], '#,##0'),
        ('Salespeople credited', tot['salespeople'], '#,##0'),
        ('Days with collection', tot['days'], '#,##0'),
        ('Average per collecting day', tot['per_day'], MONEY),
        ('First receipt', tot['first_date'], None),
        ('Last receipt', tot['last_date'], None),
    ]:
        ws.cell(row=r, column=1, value=name).border = BORDER
        c = ws.cell(row=r, column=2, value=value)
        c.border = BORDER
        if fmt:
            c.number_format = fmt
        r += 1
    _finish(ws, 'A5')

    # ---- Daily / per-dimension sheets ----
    for dim, sheet, label, width in [
        ('day', 'Daily', 'Date', 14),
        ('area', 'By Area', 'Area', 18),
        ('salesperson', 'By Salesperson', 'Salesperson', 30),
        ('customer', 'By Customer', 'Customer', 46),
        ('journal', 'By Journal', 'Journal', 30),
        ('month', 'Monthly', 'Month', 14),
    ]:
        ws = wb.create_sheet(sheet)
        titles(ws, f'{company_label} — Collections {sheet.lower()}', 7)
        _header(ws, 4, ['#', label, 'Collected', '% of Total', 'Receipts',
                        'Customers', 'Days Active', 'Last Receipt'],
                [5, width, 18, 11, 11, 11, 12, 14])
        rows = collections_data.breakdown(conn, filters, dim)
        base = tot['total'] or 1
        r = 5
        for i, row in enumerate(rows, start=1):
            vals = [i, row['label'], row['total'],
                    round(row['total'] / base * 100, 2), row['receipts'],
                    row['customers'], row['days'], row['last_date']]
            for j, v in enumerate(vals, start=1):
                c = ws.cell(row=r, column=j, value=v)
                c.border = BORDER
                if j == 1:
                    c.alignment = Alignment(horizontal='center')
                if j == 2:
                    c.alignment = Alignment(horizontal='right')
                if j == 3:
                    c.number_format, c.font = MONEY, Font(bold=True)
                if j == 4:
                    c.number_format = '0.0"%"'
            r += 1
        lbl = ws.cell(row=r, column=2, value=f'TOTAL — {len(rows)} rows')
        lbl.font, lbl.fill = Font(bold=True, size=11), TOT_FILL
        lbl.alignment = Alignment(horizontal='right')
        t = ws.cell(row=r, column=3, value=tot['total'])
        t.number_format, t.font, t.fill = MONEY, Font(bold=True, size=12), TOT_FILL
        for j in range(1, 9):
            ws.cell(row=r, column=j).border = BORDER
        _finish(ws, 'A5', f'A4:H{r - 1}')

    # ---- Receipt detail ----
    detail = collections_data.receipts(conn, filters, limit=20000)
    ws = wb.create_sheet('Receipt Detail')
    titles(ws, f'{company_label} — Receipt detail', 10)
    _header(ws, 4, ['Date', 'Receipt', 'Reference', 'Journal', 'Customer', 'Area',
                    'Invoice Settled', 'Invoice Date', 'Salesperson', 'Amount'],
            [12, 20, 38, 24, 40, 13, 20, 13, 26, 16])
    r = 5
    for row in detail['rows']:
        vals = [row['date'], row['doc'], row['ref'], row['journal'], row['customer'],
                row.get('area') or '—', row['invoice'] or '—', row['invoice_date'] or '',
                row['salesperson'], row['amount']]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.border = BORDER
            if j in (1, 8):
                c.number_format = DATE_FMT
            if j in (3, 5):
                c.alignment = Alignment(horizontal='right')
            if j == 10:
                c.number_format = MONEY
        if not row['applied']:
            ws.cell(row=r, column=9).fill = WARN_FILL
        r += 1
    lbl = ws.cell(row=r, column=9, value='TOTAL')
    lbl.font, lbl.fill = Font(bold=True), TOT_FILL
    lbl.alignment = Alignment(horizontal='right')
    t = ws.cell(row=r, column=10, value=tot['total'])
    t.number_format, t.font, t.fill = MONEY, Font(bold=True, size=12), TOT_FILL
    for j in range(1, 11):
        ws.cell(row=r, column=j).border = BORDER
    _finish(ws, 'A5', f'A4:J{r - 1}')

    return wb
