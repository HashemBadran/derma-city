"""Customers handed to a collection agency.

These are still real receivables in Odoo — the balance has not gone anywhere — but
they are no longer the team's to chase. Flagging them lets the collections
worklist show, hide, or isolate them without touching the ledger.

The flag lives in a local table that a sync never touches, exactly like
follow-ups and notes: Odoo has no field for this, and re-syncing must not lose it.

MATCHING
--------
The list arrives as a spreadsheet of names, usually truncated and spelled
inconsistently (ا/أ, ي/ى, ة/ه). A literal comparison finds almost nothing, so
names are normalised on both sides and tried exact -> prefix -> contains.

Where a name matches records in *both* companies, both are flagged: it is the
same business, and it is with the agency regardless of which ledger it sits in.
Anything that cannot be matched is reported rather than guessed at.
"""

import re
import unicodedata
from datetime import datetime

import db

SOURCE_DEFAULT = 'agency list'


def normalise(value):
    """Fold Arabic spelling variants so two writings of a name compare equal."""
    text = unicodedata.normalize('NFKC', str(value or ''))
    text = re.sub(r'[ً-ٟـ]', '', text)   # diacritics and tatweel
    for a, b in (('أ', 'ا'), ('إ', 'ا'), ('آ', 'ا'), ('ى', 'ي'),
                 ('ة', 'ه'), ('ؤ', 'و'), ('ئ', 'ي')):
        text = text.replace(a, b)
    text = re.sub(r'[^\w\s]', ' ', text)
    return re.sub(r'\s+', ' ', text).strip().lower()


# Rows that are obviously not customers — totals, legends, blank separators.
def is_noise(name):
    n = normalise(name)
    if not n or len(n) < 3:
        return True
    return bool(re.match(r'^(total|legend|المجموع|اجمالي)', n))


def read_list(path):
    """Pull (name, source, code, region) out of the agency spreadsheet.

    The header row is found by looking for a column called Customer Name rather
    than assuming a fixed row, so a re-exported file with an extra title line
    still imports.
    """
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    out = []
    for ws in wb.worksheets:
        header_row, cols = None, {}
        for r in range(1, min(ws.max_row, 15) + 1):
            values = {str(ws.cell(row=r, column=c).value or '').strip().lower(): c
                      for c in range(1, ws.max_column + 1)}
            if any('customer name' in k or 'اسم العميل' in k for k in values):
                header_row = r
                cols = values
                break
        if not header_row:
            continue

        def col(*names):
            for want in names:
                for key, idx in cols.items():
                    if want in key:
                        return idx
            return None

        c_name = col('customer name', 'اسم العميل')
        c_src, c_code, c_region = col('source'), col('customer code'), col('region')
        for r in range(header_row + 1, ws.max_row + 1):
            name = ws.cell(row=r, column=c_name).value
            if not name or is_noise(name):
                continue
            out.append({
                'name': str(name).strip(),
                'source': str(ws.cell(row=r, column=c_src).value or '').strip() if c_src else '',
                'code': str(ws.cell(row=r, column=c_code).value or '').strip() if c_code else '',
                'region': str(ws.cell(row=r, column=c_region).value or '').strip() if c_region else '',
            })
    return out


def match(conn, entries):
    """Resolve names to partner ids. Returns (hits, misses)."""
    partners = [dict(r) for r in conn.execute(
        'SELECT partner_id, name, company FROM customers WHERE partner_id > 0')]
    for p in partners:
        p['norm'] = normalise(p['name'])

    hits, misses = [], []
    for entry in entries:
        needle = normalise(entry['name'])
        if not needle:
            continue
        exact = [p for p in partners if p['norm'] == needle]
        prefix = [p for p in partners
                  if p['norm'].startswith(needle) or needle.startswith(p['norm'])]
        contains = [p for p in partners if needle in p['norm'] or p['norm'] in needle]
        found = exact or prefix or contains
        how = 'exact' if exact else ('prefix' if prefix else
                                     ('contains' if contains else ''))
        if not found:
            misses.append(entry)
            continue
        # Same business in both companies -> flag every record it resolves to.
        for p in found:
            hits.append({**entry, 'partner_id': p['partner_id'],
                         'odoo_name': p['name'], 'company': p['company'], 'how': how})
    return hits, misses


def import_file(conn, path, source=SOURCE_DEFAULT, replace=True):
    entries = read_list(path)
    hits, misses = match(conn, entries)
    stamp = datetime.now().isoformat(timespec='seconds')
    with conn:
        if replace:
            conn.execute('DELETE FROM agency WHERE source = ?', (source,))
        for h in hits:
            conn.execute(
                'INSERT INTO agency (partner_id, name, source, matched_as, added_at)'
                ' VALUES (?,?,?,?,?)'
                ' ON CONFLICT(partner_id) DO UPDATE SET name = excluded.name,'
                '   source = excluded.source, matched_as = excluded.matched_as',
                (h['partner_id'], h['odoo_name'], source, h['how'], stamp))
    return {
        'listed': len(entries),
        'flagged': len({h['partner_id'] for h in hits}),
        'unmatched': [m['name'] for m in misses],
    }


def flagged_ids(conn):
    return {r['partner_id'] for r in conn.execute('SELECT partner_id FROM agency')}


def listing(conn):
    rows = [dict(r) for r in conn.execute(
        'SELECT a.partner_id, a.name, a.source, a.matched_as, a.added_at,'
        '       c.company, c.area,'
        '       (SELECT ROUND(SUM(d.residual),2) FROM documents d'
        '         WHERE d.partner_id = a.partner_id) AS balance'
        '  FROM agency a LEFT JOIN customers c ON c.partner_id = a.partner_id'
        ' ORDER BY balance DESC')]
    for r in rows:
        r['balance'] = r['balance'] or 0.0
    return rows


def set_flag(conn, partner_id, on, source=SOURCE_DEFAULT):
    with conn:
        if on:
            row = conn.execute('SELECT name FROM customers WHERE partner_id = ?',
                               (partner_id,)).fetchone()
            conn.execute(
                'INSERT INTO agency (partner_id, name, source, matched_as, added_at)'
                ' VALUES (?,?,?,?,?) ON CONFLICT(partner_id) DO NOTHING',
                (partner_id, row['name'] if row else '', source, 'manual',
                 datetime.now().isoformat(timespec='seconds')))
        else:
            conn.execute('DELETE FROM agency WHERE partner_id = ?', (partner_id,))
    return bool(conn.execute('SELECT 1 FROM agency WHERE partner_id = ?',
                             (partner_id,)).fetchone())
